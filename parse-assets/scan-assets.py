from pathlib import Path
from collections import defaultdict
import json, re
from lxml import etree
from openpyxl import load_workbook

def ln(tag): return tag.split("}",1)[1] if "}" in tag else tag
LITS = re.compile(r'["\']([^"\']+)["\']')
def str_lits(s): return set(m.group(1).strip() for m in LITS.finditer(s or ""))

def excel_assets():
    names = {}
    for x in Path(".").rglob("Config.xlsx"):
        if ".git" in x.parts: continue
        try: wb = load_workbook(str(x), data_only=True)
        except Exception: continue
        sh = next((wb[s] for s in wb.sheetnames if s.strip().lower()=="assets"), None)
        if not sh: continue
        header = None
        for r in sh.iter_rows(min_row=1, max_row=5, values_only=True):
            if any(c not in (None,"") for c in r):
                header = [str(c).strip().lower() if c else "" for c in r]; break
        if not header: continue
        def col(name):
            for i,h in enumerate(header):
                if name in h: return i
            return -1
        i_name = col("name"); i_type = col("type")
        if i_name == -1: i_name = 0
        for r in sh.iter_rows(min_row=2, values_only=True):
            if not r or i_name>=len(r) or r[i_name] in (None,""): continue
            n = str(r[i_name]).strip()
            t = (str(r[i_type]).strip() if i_type!=-1 and i_type<len(r) and r[i_type] else "Unknown")
            names[n] = t
    return names  # {name:type}

def xaml_assets():
    direct, creds = defaultdict(set), set()
    for p in Path(".").rglob("*.xaml"):
        if ".git" in p.parts: continue
        raw = p.read_text(encoding="utf-8", errors="ignore")
        try:
            root = etree.fromstring(raw.encode("utf-8"), parser=etree.XMLParser(recover=True, huge_tree=True))
        except Exception:
            continue
        for e in root.iter():
            t = ln(e.tag)
            is_get_asset = (t=="GetAsset" or t.endswith("GetAsset"))
            is_get_cred  = (t=="GetCredential" or t.endswith("GetCredential"))
            if not (is_get_asset or is_get_cred): continue
            for k,v in e.attrib.items():
                k = ln(k)
                if k in ("AssetName","Asset","Name") or k.endswith(("AssetName","Asset")):
                    for n in str_lits(v):
                        direct[n].add(str(p))
                        if is_get_cred: creds.add(n)
            for c in e:
                k = ln(c.tag)
                if k in ("AssetName","Asset","Name") or k.endswith(("AssetName","Asset")) or ".AssetName" in k or ".Asset" in k:
                    val = (c.text or "").strip()
                    vals = {val} if val and "[" not in val and "]" not in val else str_lits(val)
                    for n in vals:
                        direct[n].add(str(p))
                        if is_get_cred: creds.add(n)
    return set(direct.keys()), creds

def main():
    excel_map = excel_assets()
    xaml_all, xaml_creds = xaml_assets()  
    credentials = {n for n,t in excel_map.items() if t.lower()=="credential"} | xaml_creds
    migratable = (set(excel_map) | xaml_all) - credentials
    out = {"migratable": sorted(migratable, key=str.casefold),
           "credentials": sorted(credentials, key=str.casefold)}
    Path("assets.json").write_text(json.dumps(out, indent=2), encoding="utf-8")
    print(json.dumps(out, indent=2))

if __name__ == "__main__":
    main()
